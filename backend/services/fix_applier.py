from __future__ import annotations
import re
import shutil

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName

from backend.models.job import Issue, AutoFixAction, IssueType

_IFERROR_GUARD = re.compile(r"\bIFERROR\s*\(|\bIFNA\s*\(", re.IGNORECASE)
_VLOOKUP_PATTERN = re.compile(
    r"^=VLOOKUP\s*\(\s*(.+?)\s*,\s*(.+?)\s*,\s*(\d+)\s*,\s*(0|FALSE)\s*\)$",
    re.IGNORECASE,
)


def _wrap_iferror(formula: str) -> str:
    inner = formula[1:]  # strip leading =
    return f'=IFERROR({inner},"")'


def _vlookup_to_xlookup(formula: str) -> str | None:
    """Convert =VLOOKUP(val, range, col_idx, 0) to =XLOOKUP(val, first_col, result_col)."""
    m = _VLOOKUP_PATTERN.match(formula.strip())
    if not m:
        return None
    lookup_val = m.group(1).strip()
    table_range = m.group(2).strip()
    col_idx = int(m.group(3))

    # Parse the range to extract first column and result column
    range_m = re.match(
        r"(?:'[^']+'|[A-Za-z0-9_]+!)?(\$?[A-Z]+)\$?(\d+):(\$?[A-Z]+)\$?(\d+)",
        table_range,
        re.IGNORECASE,
    )
    if not range_m:
        return None

    sheet_prefix_m = re.match(r"((?:'[^']+'|[A-Za-z0-9_]+)!)", table_range, re.IGNORECASE)
    sheet_prefix = sheet_prefix_m.group(1) if sheet_prefix_m else ""

    first_col = range_m.group(1).replace("$", "")
    row_start = range_m.group(2)
    last_col = range_m.group(3).replace("$", "")
    row_end = range_m.group(4)

    def col_index(c: str) -> int:
        result = 0
        for ch in c.upper():
            result = result * 26 + (ord(ch) - ord("A") + 1)
        return result

    def col_letters(idx: int) -> str:
        result = ""
        while idx > 0:
            idx, rem = divmod(idx - 1, 26)
            result = chr(65 + rem) + result
        return result

    start_idx = col_index(first_col)
    result_col = col_letters(start_idx + col_idx - 1)

    lookup_array = f"{sheet_prefix}{first_col}{row_start}:{first_col}{row_end}"
    return_array = f"{sheet_prefix}{result_col}{row_start}:{result_col}{row_end}"

    return f"=XLOOKUP({lookup_val},{lookup_array},{return_array})"


def apply_fixes(
    input_path: str,
    output_path: str,
    approved_issues: list[Issue],
    named_range_choices: dict[str, str],
) -> None:
    shutil.copy2(input_path, output_path)
    wb = load_workbook(output_path)

    for issue in approved_issues:
        sheet_name = issue.sheet_name
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Single-cell fixes
        cell_range = issue.cell_range
        # Skip if it's a multi-cell range reference for now (safety)
        if ":" in cell_range:
            continue

        try:
            cell = ws[cell_range]
        except Exception:
            continue

        action = issue.auto_fix_action

        if action == AutoFixAction.ADD_IFERROR:
            formula = cell.value
            if not isinstance(formula, str) or not formula.startswith("="):
                continue
            if _IFERROR_GUARD.search(formula):
                continue
            cell.value = _wrap_iferror(formula)

        elif action == AutoFixAction.REPLACE_VLOOKUP_XLOOKUP:
            formula = cell.value
            if not isinstance(formula, str):
                continue
            new_formula = _vlookup_to_xlookup(formula)
            if new_formula:
                cell.value = new_formula

        elif action == AutoFixAction.REPLACE_FORMULA:
            if issue.suggested_fix:
                cell.value = issue.suggested_fix

        elif action == AutoFixAction.ADD_COMMENT:
            existing = cell.comment
            note_text = f"[Excel Reviewer] {issue.description}"
            new_comment = Comment(note_text, "Excel Reviewer")
            cell.comment = new_comment

        elif action == AutoFixAction.NONE:
            # For REQUIRES_VALIDATION issues approved by the user, apply suggested_fix if present
            if issue.suggested_fix and isinstance(cell.value, str):
                cell.value = issue.suggested_fix

            # Handle named range creation for redundant_link issues
            if issue.issue_type == IssueType.REDUNDANT_LINK:
                chosen_option = named_range_choices.get(issue.issue_id)
                simp = issue.simplification_options
                if chosen_option == "b_named_range" and simp and simp.b_named_range:
                    # Extract range name from the suggestion if user didn't provide one
                    range_name = named_range_choices.get(f"{issue.issue_id}_name")
                    if range_name:
                        # Define the named range pointing to the original source
                        # The direct ref formula is the target
                        if simp.a_direct_ref:
                            ref_formula = simp.a_direct_ref.lstrip("=")
                            try:
                                defn = DefinedName(name=range_name, attr_text=ref_formula)
                                wb.defined_names.add(defn)
                                cell.value = f"={range_name}"
                            except Exception:
                                pass
                elif chosen_option == "a_direct_ref" and simp and simp.a_direct_ref:
                    cell.value = simp.a_direct_ref
                elif chosen_option == "c_structured_ref" and simp and simp.c_structured_ref:
                    if simp.c_structured_ref != "Not applicable":
                        cell.value = simp.c_structured_ref

    wb.save(output_path)
