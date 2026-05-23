from __future__ import annotations
import re

from openpyxl import load_workbook

from backend.models.analysis import CellEntry, WorkbookAnalysis
from backend.services.formula_tokenizer import is_pure_passthrough
from backend.services.graph_builder import build_dependency_graph, find_circular_references, find_relay_chains

FORMULA_ERRORS = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NAME?", "#NULL!", "#NUM!"}

VOLATILE_FUNCS = {"NOW", "TODAY", "RAND", "RANDBETWEEN", "OFFSET", "INDIRECT", "INFO", "CELL"}

_VLOOKUP_RE = re.compile(r"\bVLOOKUP\s*\(", re.IGNORECASE)
_HLOOKUP_RE = re.compile(r"\bHLOOKUP\s*\(", re.IGNORECASE)


def scan_workbook(file_path: str) -> WorkbookAnalysis:
    wb_formulas = load_workbook(file_path, data_only=False)
    wb_values = load_workbook(file_path, data_only=True)

    formula_cells: list[CellEntry] = []
    error_cells: list[CellEntry] = []
    volatile_cells: list[dict] = []
    vlookup_cells: list[CellEntry] = []
    hardcoded_candidates: list[dict] = []
    redundant_links: list[CellEntry] = []

    for sheet_idx, sheet in enumerate(wb_formulas.worksheets):
        val_sheet = wb_values.worksheets[sheet_idx]

        for row in sheet.iter_rows():
            for cell in row:
                formula = cell.value
                cached = val_sheet.cell(cell.row, cell.column).value

                if isinstance(formula, str) and formula.startswith("="):
                    cached_str = str(cached) if cached is not None else None
                    entry = CellEntry(
                        sheet=sheet.title,
                        coord=cell.coordinate,
                        formula=formula,
                        cached_value=cached_str,
                    )
                    formula_cells.append(entry)

                    if cached_str in FORMULA_ERRORS:
                        error_cells.append(entry)

                    upper = formula.upper()
                    for vf in VOLATILE_FUNCS:
                        if f"{vf}(" in upper:
                            volatile_cells.append(
                                {"sheet": sheet.title, "coord": cell.coordinate, "formula": formula, "volatile_func": vf}
                            )
                            break

                    if _VLOOKUP_RE.search(formula) or _HLOOKUP_RE.search(formula):
                        vlookup_cells.append(entry)

                    if is_pure_passthrough(formula):
                        redundant_links.append(entry)

                elif formula is not None and not isinstance(formula, str):
                    hardcoded_candidates.append(
                        {"sheet": sheet.title, "coord": cell.coordinate, "value": formula}
                    )

    G = build_dependency_graph(wb_formulas)
    cycles = find_circular_references(G)

    passthrough_ids = {f"{e.sheet}!{e.coord}" for e in redundant_links}
    relay_chains = find_relay_chains(G, passthrough_ids)

    return WorkbookAnalysis(
        sheet_names=[s.title for s in wb_formulas.worksheets],
        formula_cells=formula_cells,
        error_cells=error_cells,
        volatile_cells=volatile_cells,
        vlookup_cells=vlookup_cells,
        hardcoded_candidates=hardcoded_candidates,
        circular_references=cycles,
        redundant_links=redundant_links,
        relay_chains=relay_chains,
        total_formula_count=len(formula_cells),
        total_error_count=len(error_cells),
    )
