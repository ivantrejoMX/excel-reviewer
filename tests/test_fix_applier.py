import shutil
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from backend.models.job import Issue, IssueType, IssueSeverity, FixBucket, AutoFixAction
from backend.services.fix_applier import apply_fixes, _wrap_iferror, _vlookup_to_xlookup


# ── Unit tests for helpers ─────────────────────────────────────────────────────

def test_wrap_iferror_basic():
    assert _wrap_iferror("=A1/B1") == '=IFERROR(A1/B1,"")'


def test_wrap_iferror_keeps_equals():
    result = _wrap_iferror("=VLOOKUP(A1,B1:C5,2,0)")
    assert result.startswith("=IFERROR(")
    assert "VLOOKUP" in result


def test_vlookup_to_xlookup_basic():
    result = _vlookup_to_xlookup("=VLOOKUP(A1,B1:C10,2,0)")
    assert result is not None
    assert result.startswith("=XLOOKUP(")
    assert "A1" in result


def test_vlookup_to_xlookup_no_match():
    # Approximate match (1) should not convert
    result = _vlookup_to_xlookup("=VLOOKUP(A1,B1:C10,2,1)")
    assert result is None


def test_vlookup_to_xlookup_case_insensitive():
    result = _vlookup_to_xlookup("=vlookup(A1,B1:C10,2,FALSE)")
    assert result is not None


# ── Integration: fix_applier writes correct output ────────────────────────────

def _make_test_file(cells: dict) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for coord, val in cells.items():
        ws[coord] = val
    tmp = tempfile.mktemp(suffix=".xlsx")
    wb.save(tmp)
    return tmp


def _make_issue(issue_id, issue_type, sheet, cell, orig, fix_bucket, action, suggested=None):
    return Issue(
        issue_id=issue_id,
        issue_type=issue_type,
        severity=IssueSeverity.WARNING,
        sheet_name=sheet,
        cell_range=cell,
        description="test",
        original_formula=orig,
        suggested_fix=suggested,
        fix_bucket=fix_bucket,
        auto_fix_action=action,
    )


def test_iferror_applied():
    src = _make_test_file({"A1": "=B1/C1", "B1": 10, "C1": 0})
    out = tempfile.mktemp(suffix=".xlsx")
    issue = _make_issue(
        "i1", IssueType.FORMULA_ERROR, "Sheet1", "A1",
        "=B1/C1", FixBucket.AUTO_FIX, AutoFixAction.ADD_IFERROR,
    )
    apply_fixes(src, out, [issue], {})
    wb = load_workbook(out, data_only=False)
    assert "IFERROR" in wb["Sheet1"]["A1"].value.upper()
    Path(src).unlink(missing_ok=True)
    Path(out).unlink(missing_ok=True)


def test_replace_formula_applied():
    src = _make_test_file({"A1": "=VLOOKUP(B1,C1:D5,2,0)", "B1": "X"})
    out = tempfile.mktemp(suffix=".xlsx")
    new_formula = "=XLOOKUP(B1,C1:C5,D1:D5)"
    issue = _make_issue(
        "i2", IssueType.BROKEN_LOOKUP, "Sheet1", "A1",
        "=VLOOKUP(B1,C1:D5,2,0)", FixBucket.REQUIRES_VALIDATION,
        AutoFixAction.NONE, suggested=new_formula,
    )
    apply_fixes(src, out, [issue], {})
    wb = load_workbook(out, data_only=False)
    assert wb["Sheet1"]["A1"].value == new_formula
    Path(src).unlink(missing_ok=True)
    Path(out).unlink(missing_ok=True)


def test_no_double_iferror():
    src = _make_test_file({"A1": "=IFERROR(B1/C1,0)"})
    out = tempfile.mktemp(suffix=".xlsx")
    issue = _make_issue(
        "i3", IssueType.FORMULA_ERROR, "Sheet1", "A1",
        "=IFERROR(B1/C1,0)", FixBucket.AUTO_FIX, AutoFixAction.ADD_IFERROR,
    )
    apply_fixes(src, out, [issue], {})
    wb = load_workbook(out, data_only=False)
    # Should NOT double-wrap
    val = wb["Sheet1"]["A1"].value
    assert val.upper().count("IFERROR") == 1
    Path(src).unlink(missing_ok=True)
    Path(out).unlink(missing_ok=True)
