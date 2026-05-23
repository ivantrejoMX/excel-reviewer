"""Run this script once to generate test fixture workbooks."""
from pathlib import Path
from openpyxl import Workbook

FIXTURES = Path(__file__).parent


def make_circular_refs():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # A1 depends on B1, B1 depends on A1 → circular
    ws["A1"] = "=B1+1"
    ws["B1"] = "=A1+1"
    # Unrelated good formula
    ws["C1"] = "=SUM(D1:D5)"
    ws["D1"] = 100
    wb.save(FIXTURES / "circular_refs.xlsx")
    print("Created circular_refs.xlsx")


def make_formula_errors():
    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws["A1"] = 10
    ws["B1"] = 0
    ws["C1"] = "=A1/B1"          # #DIV/0! candidate
    ws["D1"] = "=VLOOKUP(99,A1:B5,3,0)"  # col_idx too large → #REF! candidate
    ws["E1"] = "=A1+B1"          # clean formula
    wb.save(FIXTURES / "formula_errors.xlsx")
    print("Created formula_errors.xlsx")


def make_simplification_candidates():
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    # Deeply nested IF
    ws["A1"] = '=IF(B1>100,"High",IF(B1>50,"Medium",IF(B1>10,"Low","Very Low")))'
    ws["B1"] = 75
    # VLOOKUP that could be XLOOKUP
    ws["A2"] = "=VLOOKUP(B2,D2:E10,2,0)"
    ws["B2"] = "Apple"
    ws["D2"] = "Apple"; ws["E2"] = 1.5
    ws["D3"] = "Banana"; ws["E3"] = 0.8
    # Volatile function
    ws["A3"] = "=NOW()"
    wb.save(FIXTURES / "simplification_candidates.xlsx")
    print("Created simplification_candidates.xlsx")


def make_redundant_links():
    wb = Workbook()

    # Source sheet
    src = wb.active
    src.title = "Input"
    src["C12"] = 1000
    src["C13"] = 2000

    # Relay sheet
    rel = wb.create_sheet("Assumptions")
    rel["D8"] = "=Input!C12"   # pure passthrough

    # Consumer sheet
    cons = wb.create_sheet("Summary")
    cons["B5"] = "=Assumptions!D8"  # passthrough chain: Summary→Assumptions→Input

    wb.save(FIXTURES / "redundant_links.xlsx")
    print("Created redundant_links.xlsx")


if __name__ == "__main__":
    make_circular_refs()
    make_formula_errors()
    make_simplification_candidates()
    make_redundant_links()
    print("All fixtures created.")
