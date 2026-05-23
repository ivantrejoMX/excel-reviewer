import pytest
import networkx as nx
from openpyxl import Workbook

from backend.services.graph_builder import (
    build_dependency_graph,
    find_circular_references,
    find_relay_chains,
    _iterative_has_cycle,
)


def _make_wb(cells: dict[str, str]) -> Workbook:
    """cells: {coordinate: value}"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for coord, val in cells.items():
        ws[coord] = val
    return wb


def test_no_cycle():
    wb = _make_wb({"A1": "=B1+1", "B1": "=C1*2", "C1": 5})
    G = build_dependency_graph(wb)
    assert not _iterative_has_cycle(G)
    assert find_circular_references(G) == []


def test_simple_cycle():
    wb = _make_wb({"A1": "=B1+1", "B1": "=A1+1"})
    G = build_dependency_graph(wb)
    assert _iterative_has_cycle(G)
    cycles = find_circular_references(G)
    assert len(cycles) >= 1
    # Both cells should appear in the cycle
    flat = [cell for cycle in cycles for cell in cycle]
    assert any("A1" in c for c in flat)
    assert any("B1" in c for c in flat)


def test_no_formula_cells():
    wb = _make_wb({"A1": 10, "B1": 20})
    G = build_dependency_graph(wb)
    assert G.number_of_nodes() == 0
    assert find_circular_references(G) == []


def test_relay_chains():
    # Sheet1!A1 → Sheet1!B1 → Sheet1!C1 (all passthroughs)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["C1"] = 100
    ws2 = wb.create_sheet("Sheet2")
    ws2["B1"] = "=Sheet1!C1"
    ws3 = wb.create_sheet("Sheet3")
    ws3["A1"] = "=Sheet2!B1"

    G = build_dependency_graph(wb)
    passthroughs = {"Sheet2!B1", "Sheet3!A1"}
    chains = find_relay_chains(G, passthroughs)
    assert len(chains) >= 1
    chain_flat = [c for ch in chains for c in ch]
    assert "Sheet3!A1" in chain_flat
    assert "Sheet2!B1" in chain_flat
